from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
import json
import os
import uuid
import hashlib
from datetime import datetime
from werkzeug.utils import secure_filename
import pandas as pd

app = Flask(__name__)
app.secret_key = 'profecia_dashboard_2026_secret_key'

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

USERS_FILE = os.path.join(DATA_DIR, 'users.json')
COMPANIES_FILE = os.path.join(DATA_DIR, 'companies.json')

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def init_data():
    if not os.path.exists(USERS_FILE):
        users = {
            "admin": {
                "id": "admin",
                "username": "admin",
                "password": hash_password("admin123"),
                "role": "admin",
                "name": "Administrador",
                "email": "admin@profecia.com.br",
                "created_at": datetime.now().isoformat()
            }
        }
        save_json(USERS_FILE, users)

    if not os.path.exists(COMPANIES_FILE):
        save_json(COMPANIES_FILE, {})

def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_json(filepath, data):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

def get_users():
    return load_json(USERS_FILE)

def get_companies():
    return load_json(COMPANIES_FILE)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Acesso restrito ao administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def parse_excel_financeiro(filepath):
    """Parse the financial Excel file (Financeiro type) and extract key data."""
    try:
        xl = pd.ExcelFile(filepath)
        sheets = xl.sheet_names
        data = {'sheets': sheets, 'summary': {}, 'transactions': [], 'charts': {}}

        # Try Base sheet (transaction data)
        if 'Base' in sheets:
            df = pd.read_excel(filepath, sheet_name='Base')
            df.columns = df.columns.str.strip()
            df = df.dropna(subset=['Valor'] if 'Valor' in df.columns else [df.columns[0]])
            
            if 'Valor' in df.columns and 'Tipo' in df.columns:
                entradas = df[df['Tipo'] == 'Entrada']['Valor'].sum()
                saidas = abs(df[df['Tipo'] == 'Saída']['Valor'].sum())
                saldo = entradas - saidas
                data['summary'] = {
                    'total_entradas': round(float(entradas), 2),
                    'total_saidas': round(float(saidas), 2),
                    'saldo': round(float(saldo), 2)
                }
                
                # Monthly chart data
                if 'Data' in df.columns:
                    df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
                    df['Mes'] = df['Data'].dt.to_period('M').astype(str)
                    monthly = df.groupby(['Mes', 'Tipo'])['Valor'].sum().reset_index()
                    months = sorted(df['Mes'].dropna().unique().tolist())
                    entradas_monthly = []
                    saidas_monthly = []
                    for m in months:
                        e = monthly[(monthly['Mes'] == m) & (monthly['Tipo'] == 'Entrada')]['Valor'].sum()
                        s = monthly[(monthly['Mes'] == m) & (monthly['Tipo'] == 'Saída')]['Valor'].sum()
                        entradas_monthly.append(round(float(e), 2))
                        saidas_monthly.append(round(float(abs(s)), 2))
                    data['charts']['monthly'] = {
                        'labels': months,
                        'entradas': entradas_monthly,
                        'saidas': saidas_monthly
                    }
                
                # Category breakdown
                if 'Categoria' in df.columns:
                    cat = df[df['Tipo'] == 'Saída'].groupby('Categoria')['Valor'].sum().abs()
                    cat = cat.nlargest(8)
                    data['charts']['categorias'] = {
                        'labels': cat.index.tolist(),
                        'values': [round(float(v), 2) for v in cat.values]
                    }
                
                # Recent transactions
                recent = df.sort_values('Data', ascending=False).head(20)
                data['transactions'] = []
                for _, row in recent.iterrows():
                    data['transactions'].append({
                        'data': str(row.get('Data', ''))[:10],
                        'descricao': str(row.get('Descricao', '')),
                        'categoria': str(row.get('Categoria', '')),
                        'valor': round(float(row.get('Valor', 0)), 2),
                        'tipo': str(row.get('Tipo', '')),
                        'status': str(row.get('Status', ''))
                    })

        return data
    except Exception as e:
        return {'error': str(e), 'sheets': [], 'summary': {}, 'transactions': [], 'charts': {}}

def parse_excel_profecia(filepath):
    """Parse the Profecia financial Excel file."""
    try:
        xl = pd.ExcelFile(filepath)
        sheets = xl.sheet_names
        data = {'sheets': sheets, 'summary': {}, 'transactions': [], 'charts': {}}

        # Try to get summary from DASH sheet  
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'DASH' in wb.sheetnames:
                ws = wb['DASH']
                vals = []
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            vals.append(cell)
                # Extract key values from DASH
                for i, v in enumerate(vals):
                    if 'SALDO INICIAL' in str(v):
                        if i + 1 < len(vals):
                            try:
                                data['summary']['saldo_inicial'] = float(vals[i+1])
                            except:
                                pass
                    elif 'GERAÇÃO CAIXA' in str(v):
                        if i + 1 < len(vals):
                            try:
                                data['summary']['geracao_caixa'] = float(vals[i+1])
                            except:
                                pass
                    elif 'SALDO FINAL' in str(v):
                        if i + 1 < len(vals):
                            try:
                                data['summary']['saldo_final'] = float(vals[i+1])
                            except:
                                pass
        except:
            pass

        return data
    except Exception as e:
        return {'error': str(e), 'sheets': [], 'summary': {}, 'transactions': [], 'charts': {}}

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        users = get_users()
        
        user = users.get(username)
        if user and user['password'] == hash_password(password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['name'] = user['name']
            if user['role'] == 'client':
                session['company_id'] = user.get('company_id')
            flash(f'Bem-vindo, {user["name"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuário ou senha incorretos.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role') == 'admin':
        companies = get_companies()
        users = get_users()
        client_count = sum(1 for u in users.values() if u['role'] == 'client')
        return render_template('admin_dashboard.html', 
                               companies=companies, 
                               client_count=client_count,
                               company_count=len(companies))
    else:
        # Client dashboard
        company_id = session.get('company_id')
        companies = get_companies()
        company = companies.get(company_id, {})
        
        # Load company data
        company_data = {}
        data_file = os.path.join(DATA_DIR, f'company_{company_id}_data.json')
        if os.path.exists(data_file):
            company_data = load_json(data_file)
        
        return render_template('client_dashboard.html', company=company, data=company_data)

# ─── ADMIN: COMPANIES ──────────────────────────────────────────────────────────

@app.route('/admin/companies')
@admin_required
def admin_companies():
    companies = get_companies()
    users = get_users()
    # Map companies to their client users
    company_users = {}
    for uid, u in users.items():
        if u['role'] == 'client' and u.get('company_id'):
            company_users[u['company_id']] = u
    return render_template('admin_companies.html', companies=companies, company_users=company_users)

@app.route('/admin/companies/new', methods=['GET', 'POST'])
@admin_required
def admin_new_company():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        segment = request.form.get('segment', '').strip()
        client_username = request.form.get('client_username', '').strip()
        client_password = request.form.get('client_password', '').strip()
        client_name = request.form.get('client_name', '').strip()
        client_email = request.form.get('client_email', '').strip()
        
        if not name or not client_username or not client_password:
            flash('Preencha todos os campos obrigatórios.', 'danger')
            return render_template('admin_new_company.html')
        
        users = get_users()
        if client_username in users:
            flash('Nome de usuário já existe.', 'danger')
            return render_template('admin_new_company.html')
        
        company_id = str(uuid.uuid4())[:8]
        companies = get_companies()
        companies[company_id] = {
            'id': company_id,
            'name': name,
            'cnpj': cnpj,
            'segment': segment,
            'created_at': datetime.now().isoformat(),
            'has_data': False,
            'last_upload': None
        }
        save_json(COMPANIES_FILE, companies)
        
        users[client_username] = {
            'id': client_username,
            'username': client_username,
            'password': hash_password(client_password),
            'role': 'client',
            'name': client_name or name,
            'email': client_email,
            'company_id': company_id,
            'created_at': datetime.now().isoformat()
        }
        save_json(USERS_FILE, users)
        
        flash(f'Empresa "{name}" criada com sucesso!', 'success')
        return redirect(url_for('admin_companies'))
    
    return render_template('admin_new_company.html')

@app.route('/admin/companies/<company_id>')
@admin_required
def admin_company_detail(company_id):
    companies = get_companies()
    company = companies.get(company_id)
    if not company:
        flash('Empresa não encontrada.', 'danger')
        return redirect(url_for('admin_companies'))
    
    company_data = {}
    data_file = os.path.join(DATA_DIR, f'company_{company_id}_data.json')
    if os.path.exists(data_file):
        company_data = load_json(data_file)
    
    users = get_users()
    company_user = next((u for u in users.values() if u.get('company_id') == company_id), None)
    
    return render_template('admin_company_detail.html', 
                           company=company, 
                           data=company_data,
                           company_user=company_user)

@app.route('/admin/companies/<company_id>/upload', methods=['POST'])
@admin_required
def admin_upload_excel(company_id):
    companies = get_companies()
    company = companies.get(company_id)
    if not company:
        flash('Empresa não encontrada.', 'danger')
        return redirect(url_for('admin_companies'))
    
    if 'file' not in request.files:
        flash('Nenhum arquivo selecionado.', 'danger')
        return redirect(url_for('admin_company_detail', company_id=company_id))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado.', 'danger')
        return redirect(url_for('admin_company_detail', company_id=company_id))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, f'{company_id}_{filename}')
        file.save(filepath)
        
        # Parse based on filename type
        filename_lower = filename.lower()
        if 'financeiro' in filename_lower or 'base' in filename_lower:
            parsed = parse_excel_financeiro(filepath)
        else:
            parsed = parse_excel_financeiro(filepath)
            if not parsed.get('summary'):
                parsed2 = parse_excel_profecia(filepath)
                if parsed2.get('summary'):
                    parsed = parsed2
        
        parsed['filename'] = filename
        parsed['uploaded_at'] = datetime.now().isoformat()
        parsed['company_id'] = company_id
        
        data_file = os.path.join(DATA_DIR, f'company_{company_id}_data.json')
        save_json(data_file, parsed)
        
        companies[company_id]['has_data'] = True
        companies[company_id]['last_upload'] = datetime.now().isoformat()
        save_json(COMPANIES_FILE, companies)
        
        flash(f'Excel processado com sucesso! {len(parsed.get("transactions", []))} transações importadas.', 'success')
    else:
        flash('Formato inválido. Use .xlsx ou .xls', 'danger')
    
    return redirect(url_for('admin_company_detail', company_id=company_id))

@app.route('/admin/companies/<company_id>/delete', methods=['POST'])
@admin_required
def admin_delete_company(company_id):
    companies = get_companies()
    if company_id in companies:
        company_name = companies[company_id]['name']
        del companies[company_id]
        save_json(COMPANIES_FILE, companies)
        
        # Remove associated user
        users = get_users()
        to_delete = [uid for uid, u in users.items() if u.get('company_id') == company_id]
        for uid in to_delete:
            del users[uid]
        save_json(USERS_FILE, users)
        
        # Remove data file
        data_file = os.path.join(DATA_DIR, f'company_{company_id}_data.json')
        if os.path.exists(data_file):
            os.remove(data_file)
        
        flash(f'Empresa "{company_name}" removida.', 'success')
    return redirect(url_for('admin_companies'))

# ─── ADMIN: USERS ─────────────────────────────────────────────────────────────

@app.route('/admin/users')
@admin_required
def admin_users():
    users = get_users()
    companies = get_companies()
    return render_template('admin_users.html', users=users, companies=companies)

@app.route('/admin/users/<user_id>/reset-password', methods=['POST'])
@admin_required
def admin_reset_password(user_id):
    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        flash('Nova senha inválida.', 'danger')
        return redirect(url_for('admin_users'))
    
    users = get_users()
    if user_id in users:
        users[user_id]['password'] = hash_password(new_password)
        save_json(USERS_FILE, users)
        flash('Senha alterada com sucesso.', 'success')
    return redirect(url_for('admin_users'))

if __name__ == '__main__':
    init_data()
    app.run(debug=True, port=5000)
